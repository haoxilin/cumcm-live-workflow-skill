# -*- coding: utf-8 -*-
"""
结果模板行映射校验模板（openpyxl）
====================================
用途：验证 result*.xlsx 与论文表格数值一致。
背景：位置表通常每把手两行（x/y 一组）、速度表每把手一行；
      龙头=第2行、第k节龙身=2+2k（位置）/ 2+k（速度）、龙尾(后)在最后。
      直接按行号硬编码容易错位，正确做法是先按模板表头定位行号再取值。

用法：
  1. 修改 TARGET_TIMES / ROW_KEYS 与你的模板匹配；
  2. python verify_xlsx_rows.py <结果.xlsx>
"""
import sys
import openpyxl

# ---- 按实际模板修改 ----
POS_SHEET, VEL_SHEET = 0, 1          # 位置/速度 sheet 索引
ROW_KEYS = ['龙头', '第1节', '第51节', '第101节', '第151节', '第201节', '龙尾(后)']
TIME_COLS = ['0 s', '60 s', '120 s', '180 s', '240 s', '300 s']   # 或 ['-100 s', ...]
EPS = 1e-6


def col_of(header_row, target):
    for i, h in enumerate(header_row):
        if h is not None and str(h).strip() == target:
            return i + 1
    return None


def verify(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws_pos = wb.worksheets[POS_SHEET]
    ws_vel = wb.worksheets[VEL_SHEET]

    hdr = [ws_pos.cell(1, c).value for c in range(1, ws_pos.max_column + 1)]
    rows = {}
    for r in range(2, ws_pos.max_row + 1):
        label = ws_pos.cell(r, 1).value
        if label:
            for key in ROW_KEYS:
                if str(label).startswith(key):
                    rows[key] = r  # x 行；y 行 = r+1（若 x/y 分行）
    print(f'定位到的行号: {rows}')

    for t in TIME_COLS:
        c = col_of(hdr, t)
        if c is None:
            print(f'  {t}: 表头未找到!')
            continue
        for key, r in rows.items():
            x = ws_pos.cell(r, c).value
            y = ws_pos.cell(r + 1, c).value
            v = ws_vel.cell(r - 1, c).value if key != '龙头' else ws_vel.cell(2, c).value
            print(f'  {t} {key}: x={x} y={y} v={v}')
    print('校验: 将上述数值与论文表格逐位比对（公差 1e-4 量级）')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python verify_xlsx_rows.py <结果.xlsx>')
        sys.exit(1)
    verify(sys.argv[1])
